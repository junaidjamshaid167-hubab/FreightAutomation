from openpyxl import Workbook
from openpyxl.styles import Font


class ExcelExporter:

    def export(self, filename, headers, rows):

        wb = Workbook()
        ws = wb.active

        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)

        # Write rows
        for row_index, row_data in enumerate(rows, start=2):
            for col_index, value in enumerate(row_data, start=1):
                ws.cell(row=row_index, column=col_index).value = value

        wb.save(filename)