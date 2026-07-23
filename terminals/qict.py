import json
import time
import requests
from openpyxl import load_workbook

API_URL = "https://lfs.qict.com.pk/API/api/ds/v1/Ctr-Inq"

HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept": "application/json, text/plain, */*",
    "Origin": "https://lfs.qict.com.pk",
    "Referer": "https://lfs.qict.com.pk/"
}


def get_container_info(container_no):
    payload = {
        "ctrnbr": str(container_no).strip(),
        "UserID": "Test",
        "Password": "Test",
        "Cellno": "123456"
    }

    last_error = None

    for attempt in range(3):
        try:

            response = requests.post(
                API_URL,
                data=payload,
                headers=HEADERS,
                timeout=30
            )

            response.raise_for_status()

            data = response.json()

            if isinstance(data, str):
                data = json.loads(data)

            result = {}

            for item in data:
                result[item["FIELD_NAME"]] = item["FIELD_VALUE"]

            return result

        except requests.RequestException as e:
            last_error = e
            time.sleep(1)

    raise last_error


def process_excel(
    input_file,
    output_file,
    progress_callback=None,
    log_callback=None,
    stop_callback=None
):

    wb = load_workbook(input_file)
    ws = wb.active

    headers = {}
    next_col = 2

    containers = []

    for row in range(2, ws.max_row + 1):

        value = ws.cell(row=row, column=1).value

        if value:
            containers.append((row, str(value).strip()))

    total = len(containers)

    if total == 0:

        if log_callback:
            log_callback("No containers found.")

        return

    for index, (row, container) in enumerate(containers, start=1):

        if stop_callback and stop_callback():

            if log_callback:
                log_callback("Operation cancelled by user.")

            break

        if log_callback:
            log_callback(f"[{index}/{total}] Checking {container}")

        try:

            info = get_container_info(container)

            for field, value in info.items():

                if field not in headers:

                    headers[field] = next_col
                    ws.cell(row=1, column=next_col).value = field
                    next_col += 1

                ws.cell(
                    row=row,
                    column=headers[field]
                ).value = value

        except Exception as e:

            ws.cell(row=row, column=2).value = str(e)

            if log_callback:
                log_callback(f"ERROR: {container} - {e}")

        percent = int(index / total * 100)

        if progress_callback:
            progress_callback(percent)

        if index % 20 == 0:
            wb.save(output_file)

        time.sleep(0.2)

    wb.save(output_file)

    if log_callback:
        log_callback("Completed Successfully.")