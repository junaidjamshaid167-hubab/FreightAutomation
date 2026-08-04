from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment,
)
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(
    fill_type="solid",
    start_color="1F4E78",
    end_color="1F4E78",
)
ROW_FILL_1 = PatternFill(
    fill_type="solid",
    start_color="FFFFFF",
    end_color="FFFFFF",
)

ROW_FILL_2 = PatternFill(
    fill_type="solid",
    start_color="F2F2F2",
    end_color="F2F2F2",
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF",
)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws):
    """
    Style first row.
    """

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )


def freeze_header(ws):
    ws.freeze_panes = "A2"


def add_filter(ws):
    ws.auto_filter.ref = ws.dimensions


def auto_fit_columns(ws):
    for column in ws.columns:

        length = 0
        letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    length = max(length, len(str(cell.value)))
            except Exception:
                pass

        ws.column_dimensions[letter].width = min(length + 3, 50)


def alternate_rows(ws):

    for row in range(2, ws.max_row + 1):

        fill = ROW_FILL_1 if row % 2 == 0 else ROW_FILL_2

        for cell in ws[row]:
            cell.fill = fill
            cell.border = THIN_BORDER