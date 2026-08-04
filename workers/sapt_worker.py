from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PySide6.QtCore import QObject, Signal
from openpyxl import load_workbook
import os

from terminals.sapt import SAPTTerminal


class SAPTWorker(QObject):

    progress = Signal(int)
    log = Signal(str)
    finished = Signal(str)
    error = Signal(str)

    def __init__(self, input_file, output_file):
        super().__init__()

        self.input_file = input_file
        self.output_file = output_file
        self._stop_requested = False

    def stop(self):
        self._stop_requested = True

    def run(self):

        try:
            self.log.emit("SAPT worker started.")
            self.log.emit("Reading Excel file...")

            wb = load_workbook(self.input_file)
            ws = wb.active

            containers = []

            for row in ws.iter_rows(
                min_row=2,
                min_col=1,
                max_col=1,
                values_only=True
            ):
                value = row[0]

                if value:
                    container = str(value).strip().upper()

                    if container:
                        containers.append(container)

            if not containers:
                raise Exception(
                    "No container numbers found in Excel."
                )

            total = len(containers)

            self.log.emit(
                f"Found {total} container(s)."
            )

            terminal = SAPTTerminal()

            results = []

            for index, container in enumerate(containers, start=1):

                if self._stop_requested:
                    self.log.emit("Stop requested.")
                    break

                self.log.emit(
                    f"[{index}/{total}] Processing {container}"
                )

                try:

                    self.log.emit(
                        f"Searching SAPT for {container}..."
                    )

                    info = terminal.get_container_info(
                        container
                    )

                    if info:

                        self.log.emit(
                            f"{container}: SAPT data received."
                        )

                        results.append(info)

                    else:

                        self.log.emit(
                            f"{container}: No SAPT data found."
                        )

                        results.append({
                            "Container No": container,
                            "Status": "NOT FOUND"
                        })

                except Exception as exc:

                    self.log.emit(
                        f"{container}: ERROR - {exc}"
                    )

                    results.append({
                        "Container No": container,
                        "Status": "ERROR",
                        "Error": str(exc)
                    })

                percent = int(
                    index / total * 100
                )

                self.progress.emit(percent)

            if self._stop_requested:

                self.finished.emit(
                    "SAPT processing stopped."
                )

                return

            self.save_results(results)

            self.progress.emit(100)

            self.log.emit(
                f"Output saved: {self.output_file}"
            )

            self.finished.emit(
                f"SAPT processing completed successfully.\n\n"
                f"{len(results)} container(s) processed."
            )

        except Exception as exc:

            self.error.emit(
                f"SAPT Worker Error:\n{exc}"
            )

    def save_results(self, results):

        if not results:
           return

        # --------------------------------------------------
        # Make sure output folder exists
        # --------------------------------------------------

        output_dir = os.path.dirname(
            os.path.abspath(self.output_file)
        )

        os.makedirs(
            output_dir,
            exist_ok=True
        )

        # --------------------------------------------------
        # Open original workbook
        # --------------------------------------------------

        wb = load_workbook(self.input_file)

        # --------------------------------------------------
        # Remove old SAPT Results sheet
        # --------------------------------------------------

        if "SAPT Results" in wb.sheetnames:
            del wb["SAPT Results"]

        result_ws = wb.create_sheet("SAPT Results")

        # --------------------------------------------------
        # Fixed column order
        # --------------------------------------------------

        columns = [
            "Container No",
            "Status",
            "Owner",
            "BL/ Shipping Bill No.",
            "Container Size/Type",
            "Category",
            "Vessel Voyage",
            "VIR No",
            "ETA",
            "ETD",
            "Gate In Time",
            "Gate Out Time",
            "Discharge Time",
            "Load Time",
            "DO Issuance Date",
            "DO Expiry Date",
            "Origin",
            "Destination",
            "Custom Seal No.",
            "Line Seal No.",
            "Security Seal No.",
            "Other Seal No.",
            "Custom Status",
            "Current Position",
            "Commodity",
            "Weight",
            "Weighment",
            "Scanning",
            "Present Holds",
            "pid",
            "Error",
        ]

        # --------------------------------------------------
        # Header
        # --------------------------------------------------

        for col_index, column in enumerate(
            columns,
            start=1
        ):

            cell = result_ws.cell(
                row=1,
                column=col_index,
                value=column
            )

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        # --------------------------------------------------
        # Data
        # --------------------------------------------------

        for row_index, result in enumerate(
            results,
            start=2
        ):

            for col_index, column in enumerate(
                columns,
                start=1
            ):

                value = result.get(
                    column,
                    ""
                )

                cell = result_ws.cell(
                    row=row_index,
                    column=col_index,
                    value=value
                )

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )

        # --------------------------------------------------
        # Freeze header row
        # --------------------------------------------------

        result_ws.freeze_panes = "A2"

        # --------------------------------------------------
        # Enable AutoFilter
        # --------------------------------------------------

        last_column = get_column_letter(
            len(columns)
        )

        last_row = result_ws.max_row

        result_ws.auto_filter.ref = (
            f"A1:{last_column}{last_row}"
        )

        # --------------------------------------------------
        # Header row height
        # --------------------------------------------------

        result_ws.row_dimensions[1].height = 35

        # --------------------------------------------------
        # Set column widths
        # --------------------------------------------------

        widths = {
            "A": 18,   # Container No
            "B": 14,   # Status
            "C": 12,   # Owner
            "D": 25,   # BL
            "E": 20,   # Size/Type
            "F": 14,   # Category
            "G": 30,   # Vessel
            "H": 25,   # VIR
            "I": 24,   # ETA
            "J": 24,   # ETD
            "K": 24,   # Gate In
            "L": 24,   # Gate Out
            "M": 24,   # Discharge
            "N": 24,   # Load
            "O": 24,   # DO Issuance
            "P": 24,   # DO Expiry
            "Q": 14,   # Origin
            "R": 16,   # Destination
            "S": 20,   # Custom Seal
            "T": 20,   # Line Seal
            "U": 20,   # Security Seal
            "V": 18,   # Other Seal
            "W": 18,   # Custom Status
            "X": 25,   # Current Position
            "Y": 25,   # Commodity
            "Z": 14,   # Weight
            "AA": 16,  # Weighment
            "AB": 16,  # Scanning
            "AC": 18,  # Holds
            "AD": 14,  # PID
            "AE": 35,  # Error
        }

        for column, width in widths.items():

            result_ws.column_dimensions[
                column
            ].width = width

        # --------------------------------------------------
        # Borders
        # --------------------------------------------------

        thin = Side(
            style="thin"
        )

        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

        for row in result_ws.iter_rows():

            for cell in row:

                cell.border = border

        # --------------------------------------------------
        # Status formatting
        # --------------------------------------------------

        for row in range(
            2,
            result_ws.max_row + 1
        ):

            status_cell = result_ws.cell(
                row=row,
                column=2
            )

            status = str(
                status_cell.value or ""
            ).upper()

            if status == "NOT FOUND":

                status_cell.font = Font(
                    bold=True
                )

            elif status == "ERROR":

                status_cell.font = Font(
                    bold=True
                )

        # --------------------------------------------------
        # Sheet view
        # --------------------------------------------------

        result_ws.sheet_view.zoomScale = 85

        # --------------------------------------------------
        # Save
        # --------------------------------------------------

        wb.save(self.output_file)